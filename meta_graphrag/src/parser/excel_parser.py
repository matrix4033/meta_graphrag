"""Excel解析器 - 解析元数据文档"""
import openpyxl
from pathlib import Path
from typing import Dict, List, Any, Optional
from utils.logger import setup_logger


class ExcelParser:
    """Excel元数据文档解析器"""
    
    def __init__(self, config: Dict[str, Any]):
        """
        初始化Excel解析器
        
        Args:
            config: 解析器配置
        """
        self.config = config
        self.logger = setup_logger('excel_parser')
        
        # 获取列名映射配置
        self.column_mapping = config.get('column_mapping', {})
    
    def _find_column_index(self, headers: List[str], patterns: List[str]) -> Optional[int]:
        """
        根据模式查找列索引
        
        Args:
            headers: 列名列表
            patterns: 匹配模式列表
            
        Returns:
            列索引，如果未找到返回None
        """
        for i, header in enumerate(headers):
            if header:
                for pattern in patterns:
                    if pattern in header:
                        return i
        return None
    
    def parse_excel(self, excel_path: Path) -> List[Dict[str, Any]]:
        """
        解析Excel文档
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            元数据记录列表
        """
        try:
            # 加载工作簿
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # 读取表头
            headers = [cell.value for cell in ws[1]]
            
            # 查找关键列的索引
            column_indices = {}
            
            # 业务域列
            business_domain_idx = self._find_column_index(
                headers,
                self.column_mapping.get('business_domain', {}).get('patterns', ['业务域'])
            )
            
            # 一级类目（业务域）
            level1_idx = self._find_column_index(
                headers,
                self.column_mapping.get('level1', {}).get('patterns', ['一级类目'])
            )
            
            # 二级类目（业务主题）
            level2_idx = self._find_column_index(
                headers,
                self.column_mapping.get('level2', {}).get('patterns', ['二级类目'])
            )
            
            # 三级类目（逻辑实体）
            level3_idx = self._find_column_index(
                headers,
                self.column_mapping.get('level3', {}).get('patterns', ['三级类目'])
            )
            
            # 数据元名称（字段名对应的业务术语）
            data_element_idx = self._find_column_index(
                headers,
                self.column_mapping.get('data_element', {}).get('patterns', ['数据元名称'])
            )
            
            # 业务说明
            business_desc_idx = self._find_column_index(
                headers,
                self.column_mapping.get('business_description', {}).get('patterns', ['业务说明'])
            )
            
            # 数源单位
            data_source_idx = self._find_column_index(
                headers,
                self.column_mapping.get('data_source_unit', {}).get('patterns', ['数源单位'])
            )
            
            # 数据标准
            data_standard_idx = self._find_column_index(
                headers,
                self.column_mapping.get('data_standard', {}).get('patterns', ['数据标准'])
            )
            
            # 数据元格式
            data_format_idx = self._find_column_index(
                headers,
                self.column_mapping.get('data_format', {}).get('patterns', ['数据元格式'])
            )
            
            # 值域
            value_domain_idx = self._find_column_index(
                headers,
                self.column_mapping.get('value_domain', {}).get('patterns', ['值域'])
            )
            
            self.logger.info(f"找到列索引: 一级类目={level1_idx}, 二级类目={level2_idx}, "
                           f"三级类目={level3_idx}, 数据元名称={data_element_idx}, "
                           f"数据元格式={data_format_idx}, 值域={value_domain_idx}")
            
            # 解析数据行
            records = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                try:
                    # 提取业务分类
                    business_domain = row[level1_idx] if level1_idx is not None else None
                    business_subject = row[level2_idx] if level2_idx is not None else None
                    logical_entity = row[level3_idx] if level3_idx is not None else None
                    
                    # 提取数据元信息
                    data_element_name = row[data_element_idx] if data_element_idx is not None else None
                    business_description = row[business_desc_idx] if business_desc_idx is not None else None
                    data_source_unit = row[data_source_idx] if data_source_idx is not None else None
                    data_standard = row[data_standard_idx] if data_standard_idx is not None else None
                    data_format = row[data_format_idx] if data_format_idx is not None else None
                    value_domain = row[value_domain_idx] if value_domain_idx is not None else None
                    
                    # 跳过空行
                    if not logical_entity or not data_element_name:
                        continue
                    
                    record = {
                        'business_domain': business_domain,
                        'business_subject': business_subject,
                        'logical_entity': logical_entity,
                        'business_term': data_element_name,
                        'business_description': business_description,
                        'data_source_unit': data_source_unit,
                        'data_standard': data_standard,
                        'data_format': data_format,
                        'value_domain': value_domain
                    }
                    
                    records.append(record)
                    
                except Exception as e:
                    self.logger.warning(f"解析行{row_idx}失败: {e}")
                    continue
            
            self.logger.info(f"从 {excel_path} 解析了 {len(records)} 条记录")
            return records
            
        except Exception as e:
            self.logger.error(f"解析Excel文件失败: {excel_path}, 错误: {e}")
            return []
